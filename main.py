from asyncore import write
import os,time
from selenium import webdriver
from selenium.webdriver.firefox.options import Options as firefoxOptions
from selenium.webdriver.chrome.options import Options as chromeOptions
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.firefox_binary import FirefoxBinary
from webdriver_manager.firefox import GeckoDriverManager
from webdriver_manager.chrome import ChromeDriverManager
from datetime import datetime
from openpyxl import Workbook,load_workbook
from mysql import connector
from mysql.connector import Error

def func_writeDataToMySQL(big_data):
    try:
        db = connector.connect(user="FollowiZ",password="@followiz_@2022DBpw!",database="followiz_db",host="localhost")
        if db.is_connected():
            db_info = db.get_server_info()
            print("CONNECTED",db_info)
            cursor = db.cursor(buffered=True)
            cursor.execute("CREATE TABLE IF NOT EXISTS updates_new (id INT AUTO_INCREMENT PRIMARY KEY,service TEXT,date TEXT,status TEXT,update_status TEXT,created_at TIMESTAMP) DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci")
            db.commit()
            for data in big_data:
                for n in data:
                    cursor.execute("select id from updates_new where service=%s and date =%s and status=%s and update_status=%s",(n[0],n[1],n[2],n[3]))
                    service_data = cursor.fetchone()
                    #print(service_data)
                    #if service_data:
                    #    cursor.execute("update updates_new set service=%s,date=%s,status=%s,update_status=%s,created_at=%s where ID=%s",(n[0],n[1],n[2],n[3],datetime.now(),service_data[0]))
                    #    db.commit()
                    #else:
                    if service_data is None:
                        cursor.execute("insert into updates_new(service,date,status,update_status,created_at) values(%s,%s,%s,%s,%s)",(n[0],n[1],n[2],n[3],datetime.now()))
                        db.commit()
        else:
            print("CONNECTING FAILED")
    except Error as e:
        print(e)
    finally:
        db.close()

def func_login(driver):
    inputUser = driver.find_element(By.XPATH,"/html/body/div/div/div/div/div/div/form/div[1]/input")
    inputPassword = driver.find_element(By.XPATH,"/html/body/div/div/div/div/div/div/form/div[2]/input")
    confirmButton = driver.find_element(By.XPATH,"/html/body/div/div/div/div/div/div/form/button")
    inputUser.send_keys("khamid")
    inputPassword.send_keys("4bZft5cs")
    time.sleep(0.5)
    confirmButton.click()
    time.sleep(0.5)

def func_scrapDataFromPage(driver):
    tbody = driver.find_element(By.TAG_NAME,"tbody")
    all_tr = tbody.find_elements(By.TAG_NAME,"tr")
    tr_list = []
    for tr in all_tr:
        count = 0
        all_td = tr.find_elements(By.TAG_NAME,"td")
        for td in all_td:
            count += 1
            if count == 1:
                service = td.text
            elif count == 2:
                date = td.text
            elif count == 3:
                update = td.text
        update_status = tr.find_element(By.TAG_NAME,"span")
        tr_list.append((service,date,update,update_status.get_attribute("class")))
    #print(tr_list)
    return tr_list

def func_writeDataToExcel(data,path):
    try:
        xl = load_workbook(path+'/followiz.xlsx')
        xs = Workbook.active
        sheet = xl["Main"]
    except:
        xl = Workbook()
        xs = Workbook.active
        sheet = xl.create_sheet("Main")
        sheet.append(["SERVICE","DATE","UPDATE","UPDATE-STATUS"])
    for n in data:
        for col in n:
            sheet.append(col)
    try:
        os.remove(path+'/followiz.xlsx')
    except:
        pass
    xl.save(path+'/followiz.xlsx')

def scrapAllDataByZero():
    options = chromeOptions()
    #options = firefoxOptions()
    #options.binary_location = '/usr/lib/chromium-browser/chromedriver'
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--log-level=3")
    options.add_argument("--no-sandbox")
    options.add_argument('--disable-dev-shm-usage')
    #options.add_experimental_option("excludeSwitches", ["enable-logging"])
    #options.add_argument("--remote-debugging-port=9222")
    #path = FirefoxBinary.which()
    #driver = webdriver.Firefox(service= Service(executable_path=GeckoDriverManager().install()),options=options)
    #driver = webdriver.Chrome(executable_path=ChromeDriverManager().install(),options=options)
    driver = webdriver.Chrome(executable_path="/root/chromedriver",options=options)
    driver.get('https://followiz.com/updates')
    finally_list = []
    func_login(driver)
    finally_list.append(func_scrapDataFromPage(driver))
    path = os.path.dirname(os.path.realpath(__name__))
    for n in range(2,10000):
        try:
            #print(n)
            driver.get('https://followiz.com/updates/'+str(n))
            page_list = func_scrapDataFromPage(driver)
            finally_list.append(page_list)
        except:
            break
        finally:
            if n == 2:
                break
    func_writeDataToMySQL(finally_list)
    #func_writeDataToExcel(finally_list,path)
#func_writeDataToMySQL(5)
scrapAllDataByZero()
print("FUNCTION HAS DONE",datetime.now())

    

